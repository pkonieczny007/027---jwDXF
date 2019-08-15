import os 
import fdxf
import openpyxl 

n=0
index_dict={}
idi= index_dict
test_dict = {}


wb = openpyxl.Workbook()    #utworzenie nowego pliku excel
ws = wb.active

for filename in os.listdir():
    if filename.endswith('.dxf'):
        n+=1
        wys=fdxf.wys(filename)
        szer=fdxf.sze(filename)
        #print(n, filename, wys, szer)
        test_dict[filename]=(wys,szer)
        index_dict[n]=filename
        a='A'+str(n)
        b='B'+str(n)
        c='C'+str(n)
        ws[a]=filename
        ws[b]=wys
        ws[c]=szer
    else:
        print('dane zaladowane')





    



wb.save('testing_tech.xlsx')
print('plik testowy zapisany;')


        
